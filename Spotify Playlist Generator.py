import openpyxl
import streamlit as st
import spotipy
from spotipy.oauth2 import SpotifyOAuth
import logging

def playlist(playlist_name):
    logging.basicConfig(filename='spotify_playlist.log', level=logging.DEBUG,
                        format='%(asctime)s %(levelname)s %(message)s')
    wb = openpyxl.load_workbook('preferences.xlsx')
    ws = wb.active
    music_choices = [cell.value for cell in ws['A']]

    sp = spotipy.Spotify(auth_manager=SpotifyOAuth(client_id='f7a0b873ee934f5481e135cea61d1063',
                                                   client_secret='b9b6caeaf3d94992ab282d99e4934f89',
                                                   redirect_uri='http://localhost:8888/callback',
                                                   scope='playlist-modify-public'))

    try:
        playlist_description = "A playlist generated with your favourite artists' top songs"
        user_id = sp.current_user()['id']
        playlist = sp.user_playlist_create(user=user_id, name=playlist_name, public=True,
                                           description=playlist_description)
        playlist_link = playlist['external_urls']['spotify']
        st.write('Playlist created: ' + playlist_name)
        logging.info('Playlist created: ' + playlist_name)

    except Exception as e:
        print('Error creating playlist: ' + str(e))
        logging.error('Error creating playlist: ' + str(e))

    for song in music_choices:
        try:
            results = sp.search(q='artist:' + song, type='track', limit=10)
            if results['tracks']['items']:
                for item in results['tracks']['items']:
                    track_uri = item['uri']
                    artist_name = item['artists'][0]['name']
                    try:
                        sp.playlist_add_items(playlist_id=playlist['id'], items=[track_uri])
                        logging.info('Added ' + song + ' by ' + artist_name + ' to playlist')
                    except Exception as e:
                        logging.error('Error adding ' + song + ' by ' + artist_name + ' to playlist: ' + str(e))
            else:
                logging.warning('Could not find ' + song + ' in Spotify')
        except Exception as e:
            logging.error('Error searching for ' + song + ': ' + str(e))
    sheet.delete_cols(1, 1000)
    wb.save("preferences.xlsx")
    return playlist_link

#main
wb= openpyxl.load_workbook('preferences.xlsx')
sheet = wb.active
l1=[]
y=1
image = open("C:\\Users\\Soham\\Downloads\\Spotify_logo_with_text.svg.png", "rb").read()
st.set_page_config(page_title='Spotify Playlist Generator', page_icon=':notes:', layout='centered', initial_sidebar_state='collapsed')
st.image(image,width=300)
st.title("Spotify Playlist Generator")
css = """
body {
    background-color: #191414;
    color: #fff;
    font-family: 'Helvetica Neue', sans-serif;
}
.stButton button {
    background-color: #1db954;
    border-radius: 50px;
    border: none;
    color: #fff;
    font-size: 1rem;
    font-weight: 700;
    padding: 12px 24px;
    transition: all 0.2s;
}
.stButton button:hover {
    background-color: #1ed760;
    cursor: pointer;
}
"""
st.markdown(f'<style>{css}</style>', unsafe_allow_html=True)
create_type = st.radio('Do you want to Create a Playlist?', ['Yes', 'No'])
if create_type == 'Yes':
    number = st.slider('Enter The Number of Artists', value=0)
    playlist_name = st.text_input("Enter Playlist Name:")
    for i in range(number):
        name=st.text_input("Enter Name of Artist #{}".format(i+1))
        l1.append(name)
    for j in l1:
            c1 = sheet.cell(row=y, column=1)
            c1.value = j
            wb.save("preferences.xlsx")
            y=y+1
    if st.button('Create'):
        x=playlist(playlist_name)
        st.write("Playlist Link: '{}'".format(x))
else:
    st.write("Thank You Have a Nice Day!")
    sheet.delete_cols(1, 1000)
    wb.save("preferences.xlsx")
